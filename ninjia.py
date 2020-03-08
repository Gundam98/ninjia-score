# 引入load_workbook
from openpyxl import load_workbook
from aip import AipOcr
from fuzzywuzzy import fuzz

config = {
    'appId': '18723556',
    'apiKey': 'YmQ7Qi2Kooh5qOoENIAP1V4O',
    'secretKey': 'vtk5zbBVi4hoVWclNwedp1orfxjtfkOM'
}

client = AipOcr(**config)

def get_file_content(file):
    with open(file, 'rb') as fp:
        return fp.read()

def img_to_str(image_path):
    image = get_file_content(image_path)
    result = client.basicGeneral(image)
    data=[]
    if 'words_result' in result:
        for w in result['words_result']:
            data.append(w['words'])
    return data

def getImgInfo():
    imagepath = 'pic.jpg'
    rawResult = img_to_str(imagepath)
    resultList = []
    result = {}
    for word in rawResult :
        if word.find('新人') < 0 \
            and word.find('族员') < 0 \
            and word.find('精英') < 0 \
            and word.find('豪杰') < 0 \
            and word.find('长老') < 0 \
            and word.find('副族长') < 0 \
            and word.find('族长') <0:
            resultList.append(word)
    for i in range(0, len(resultList) - 1, 2) :
        result[resultList[i]] = resultList[i+1]
    return result



def writeData(result):
    # print(result)
    xlsx = load_workbook('score.xlsx')
    FuBen = xlsx['副本']
    JiaZuZhan = xlsx['家族战33']
    ShenYuan = xlsx['深渊']

    nameCells = FuBen['B']
    nameList = []
    for cell in nameCells :
        if (cell.value != 'id') and (cell.value != None) and cell.fill.start_color.index != 1 :
            nameList.append(str(cell.value))
    # print(names)

    for name, score in result.items() :
        # print('name:', name, ' score:', score)
        pos = nameList.index(name) if (name in nameList) else -1
        if pos >= 0 :
            print('姓名：', name, '\t分数：', score, '\t行数', pos + 2)
        else :
            maxSame = 0
            for i in range(len(nameList) - 1) :
                currentSame = fuzz.partial_ratio(nameList[i], name)
                if currentSame > maxSame :
                    maxSame = currentSame
                    pos = i
            print('姓名：', name, '\t最相似姓名：', nameList[pos], '\t分数：', score, '\t行数：', i + 2, '\t匹配度：', maxSame)

if __name__ == '__main__' :
    result = getImgInfo()
    writeData(result)



