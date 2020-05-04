# 操作excel的库
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Border,Side
# 获取文本匹配度的库
from fuzzywuzzy import fuzz

from ocr import getImgInfo
import utils
import globalValue as glob


dungeonList = {
    '1': '蝙蝠',
    '2': '西瓜',
    '3': '金币',
    '4': '飞镖',
    '5': '礼带',
    '6': '河豚',
    '7': '宝箱',
    '8': '逃离森林',
}

classifyJson = {
    "color": {
        "A": "4EAC5B",
        "B": "A0CD62",
        "C": "FFFD55",
        "D": "F7C143",
        "absence": "EA3323",
        "leave": "BFBFBF"
    },
    "蝙蝠": {
        "A": 390,
        "B": 385,
        "C": 380,
        "max": 401
    },
    "西瓜": {
        "A": 317,
        "B": 314,
        "C": 310,
        "max":317
    },
    "金币": {
        "A": 3090,
        "B": 3080,
        "C": 3070,
        "max":3113
    },
    "飞镖": {
        "A": 600,
        "B": 598,
        "C": 596,
        "max": 601,
    },
    "礼带": {
        "A": 76,
        "B": 73,
        "C": 70,
        "max": 82
    },
    "河豚": {
        "A": 185,
        "B": 180,
        "C": 175,
        "max": 197
    },
    "宝箱": {
        "A": 60,
        "B": 58,
        "C": 57,
        "max": 62
    },
    "逃离森林": {
        "A": 2400,
        "B": 2200,
        "C": 2000,
        "max": 2440
    },
}

nameList = []

def writeDungeonData(OCRResult, dungeonType):
    global nameList
    nameList = utils.getNameList('副本')
    xlsx = load_workbook(glob.excelPath)
    sheet = xlsx['副本']

    col = sheet.max_column + 1
    inputCol = input('将在表格第%s列插入数据。输入数字更改，或enter确认:' %(col))
    if inputCol != '' : col = int(inputCol)
    maxScore = int(classifyJson[dungeonList[dungeonType]]['max'])
    count = 0
    preScore = {'score':maxScore}
    nextScore = {'score':maxScore}
    currentScore = {'score':maxScore}
    currentRow = 0
    nextRow = 0
    for name, score in OCRResult.items():
        # print(preScore,currentScore,nextScore)
        row = nameList.index(name) if (name in nameList) else -1
        if row < 0 :
            maxSame = 0
            for i in range(len(nameList) - 1) :
                currentSame = fuzz.partial_ratio(nameList[i], name)
                if currentSame > maxSame :
                    maxSame = currentSame
                    row = i
            if row < 0 :
                print("\033[0;37;41mERROR\033[0m 没有找到族员:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认他是否改名。不然就是程序出错了诶嘿😛" % (name,score,count+1))
                continue
            elif  maxSame < 0.5:
                print("\033[0;30;43mWARN\033[0m 没有找到族员:\033[0;30;47m%s\033[0m，名字最接近的族员是:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s\033[0m。请留意匹配是否出错。" % (name,nameList[row],score))
        try:
            if 'score' in currentScore and 'score' in preScore and 'score' in nextScore and (float(currentScore['score']) > float(preScore['score']) or float(currentScore['score']) < float(nextScore['score']) or float(currentScore['score']) / float(preScore['score']) < 0.5 ):
                if 'name' in currentScore:
                    print("\033[0;30;43mINFO\033[0m 族员:\033[0;30;47m%s\033[0m的成绩较为异常。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认是否识别有误。(not the right rank)" % (nameList[currentRow],currentScore['score'],count-1))
            else:
                preScore = currentScore
            currentScore = nextScore
            currentRow = nextRow
            nextScore = {'name':name,'score':score}
            nextRow = row
        except Exception:
            pass

        score = 0
        row = 0
        try:
            if 'score' in currentScore:
                score = float(currentScore['score'])
                row = currentRow + 3
            elif  'score' in nextScore:
                score = float(nextScore['score'])
                row = nextRow + 3
            else:
                print("\033[0;37;41mERROR\033[0m。程序出错了。(get current score or next score error)")
        except Exception:
            if 'score' in currentScore:
                score = currentScore['score']
                row = currentRow + 3
            elif 'score' in nextScore:
                score = nextScore['score']
                row = nextRow + 3
            else:
                print("\033[0;37;41mERROR\033[0m。程序出错了。(get current score or next score error)")

        sheet.cell(row,col).value = score
        count += 1
    #将最后一个加进去
    try:
        if float(nextScore['score']) / float(currentScore['score']) < 0.5:
            print("\033[0;30;43mINFO\033[0m 族员:\033[0;30;47m%s\033[0m的成绩较为异常。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认是否识别有误。(not the right rank)" % (nameList[nextRow],nextScore['score'],count))
        score = float(nextScore['score'])
    except Exception:
        score = nextScore['score']
        
    row = nextRow + 3
    sheet.cell(row,col).value = score
            
    xlsx.save(glob.excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 成功录入%d条数据！" % (count))
    return col

def decorateDungeonData(col, dungeonType):
    input("输入任意键开始填充颜色……")
    xlsx = load_workbook(glob.excelPath)
    FuBen = xlsx['副本']
    classify = classifyJson[dungeonList[dungeonType]]
    colorList = classifyJson['color']
    FuBen.cell(2, col).value = dungeonList[dungeonType]

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin',color='000000'), bottom=Side(border_style='thin',color='000000'))
        
    defaultAbsence = False
    for i in range(3, len(nameList)+ 3): 
        color = ""
        if FuBen.cell(i,col).value == None:
            if not defaultAbsence:
                noScore = input("%s没有成绩，是否请假？(默认否,全部未请假输入all)[N/y/all]:" %(FuBen.cell(i,2).value))
                if noScore == '': noScore = "N"
                while noScore != 'N' and noScore != 'n' and noScore != 'Y' and noScore != 'y' and noScore != 'all':
                    noScore = input("输入无效，请输入‘y’、‘n’、‘all’或直接enter:")
                    if noScore == '': noScore = "N"
                if noScore == 'all': 
                    defaultAbsence = True
                    noScore = 'N'
            else:
                noScore = 'N'       
            color = colorList['absence'] if noScore == "N" or noScore == "n" else colorList['leave']
        elif int(FuBen.cell(i,col).value) >= classify['A']:
            color = colorList['A']
        elif int(FuBen.cell(i,col).value) < classify['A'] and int(FuBen.cell(i,col).value) >= classify['B']:
            color = colorList['B']
        elif  int(FuBen.cell(i,col).value) < classify['B'] and int(FuBen.cell(i,col).value) >= classify['C']:
            color = colorList['C']
        else:
            color = colorList['D']
        FuBen.cell(i, col).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
        FuBen.cell(i, col).border = border
        
    xlsx.save(glob.excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 颜色填充成功！")

def familyWarDungeon():
    dungeonType = input('请选择副本类型 (默认为: 礼带)\n1.蝙蝠 2.西瓜 3.金币 4.飞镖 5.礼带 6.河豚 7.宝箱 8.逃离森林\n:')
    if dungeonType == '': dungeonType = '5'
    result = getImgInfo(3)
    col = writeDungeonData(result, dungeonType)
    decorateDungeonData(col, dungeonType)

def dungeonPreparation():
    dungeonType = input('请选择副本类型 (默认为: 礼带)\n1.蝙蝠 2.西瓜 3.金币 4.飞镖 5.礼带 6.河豚 7.宝箱 8.无尽\n:')
    if dungeonType == '': dungeonType = '5'
    result = getImgInfo(1)
    col = writeDungeonData(result, dungeonType)
    decorateDungeonData(col, dungeonType)


