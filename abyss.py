# 操作excel的库
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Border,Side
# 获取文本匹配度的库
from fuzzywuzzy import fuzz

from ocr import getImgInfo
import utils
import globalValue as glob

nameList = []

def writeAbyssData(OCRResult):
    global nameList
    nameList = utils.getNameList('深渊')
    xlsx = load_workbook(glob.excelPath)
    sheet = xlsx['深渊']

    col = sheet.max_column + 1
    inputCol = input('将在表格第%s,%s,%s列插入数据。输入数字(起始列编号)更改，或enter确认:' %(col, col + 1, col + 2))
    if inputCol != '' : col = int(inputCol)

    count = 0
    for name, data in OCRResult.items() :
        damage = data[0]
        times = data[1]
        average = data[2]
        row = nameList.index(name) if (name in nameList) else -1
        print('find ' + name + 'in row ' + str(row + 3))
        if row < 0 :
            maxSame = 0
            for i in range(len(nameList) - 1) :
                currentSame = fuzz.partial_ratio(nameList[i], name)
                if currentSame > maxSame :
                    maxSame = currentSame
                    row = i
            if row < 0 :
                print("\033[0;37;41mERROR\033[0m 没有找到族员:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s,%s,%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认他是否改名。不然就是程序出错了诶嘿😛" % (name, damage, times, average, count + 1))
                continue
            elif maxSame < 0.5:
                print("\033[0;30;43mWARN\033[0m 没有找到族员:\033[0;30;47m%s\033[0m，名字最接近的族员是:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s,%s,%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请留意匹配是否出错。" % (name,nameList[row],damage, times, average, count + 1))
                            
        row += 3
        try:
            sheet.cell(row,col).value = float(damage)
            sheet.cell(row,col+1).value = float(times)
            sheet.cell(row,col+2).value = float(average)
        except Exception as e:
            print('\033[0;37;41mERROR\033[0m \033[0;30;47m%s\033[0m的成绩登记有误，请留意是否出错，他的成绩为:%s,%s,%s。他的大概排名为:\033[0;37;44m%d\033[0m。(convert to number error)' % (name, damage, times, average, count + 1))
            sheet.cell(row,col).value = damage
            sheet.cell(row,col+1).value = times
            sheet.cell(row,col+2).value = average

        count += 1
            
    xlsx.save(glob.excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 成功录入%d条数据！" % (count))
    return col

def decorateAbyssData(col):
    xlsx = load_workbook(glob.excelPath)
    sheet = xlsx['深渊']

    absenceColor = 'EA3323'
    leaveColor = 'BFBFBF'

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin',color='000000'), bottom=Side(border_style='thin',color='000000'))

    defaultAbsence = False
    for i in range(3, len(nameList)+ 3): 
        color = ""
        if sheet.cell(i,col).value == None:
            if not defaultAbsence:
                noScore = input("%s没有成绩，是否请假？(默认否,全部未请假输入all)[N/y/all]:" %(sheet.cell(i,2).value))
                if noScore == '': noScore = "N"
                while noScore != 'N' and noScore != 'n' and noScore != 'Y' and noScore != 'y' and noScore != 'all':
                    noScore = input("输入无效，请输入‘y’、‘n’、‘all’或直接enter:")
                    if noScore == '': noScore = "N"
                if noScore == 'all': 
                    defaultAbsence = True
                    noScore = 'N'
            else:
                noScore = 'N'       
            color = absenceColor if noScore == "N" or noScore == "n" else leaveColor

            sheet.cell(i, col).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
            sheet.cell(i, col+1).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
            sheet.cell(i, col+2).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
            
        #描边
        sheet.cell(i, col).border = border
        sheet.cell(i, col+1).border = border
        sheet.cell(i, col+2).border = border

    xlsx.save(glob.excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 颜色填充成功！")

def abyss():
    result = getImgInfo(2)
    col = writeAbyssData(result)
    decorateAbyssData(col)
    return



