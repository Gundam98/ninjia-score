# -*- coding:UTF-8 -*-
# -*- encoding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Border,Side
from fuzzywuzzy import fuzz
import json
import os
import sys
import base64
from urllib.parse import quote
import requests

'''
getImgInfo()的返回值的数据格式
副本:
    {
        name(string): score(string)
    }
深渊:
    {
        name(string): [damage(string), times(string), average(string)]
    }
33:
    {
        name(string): [win(boolean), rate(string)]
    }
'''


templateSignList = {
    1: '1f7f73f7e9b25b4918e02974a6683ce3',
    2: '95952d339513cf56bb46347ea14a561b',
    3: 'd833009061f3d2260944a841b4cf9f32',
    4: '',
}


dungeonList = {
    '1': '蝙蝠',
    '2': '西瓜',
    '3': '金币',
    '4': '飞镖',
    '5': '礼带',
    '6': '河豚',
    '7': '宝箱',
    '8': '无尽'
}


nameList = []


excelPath = '胧月成绩统计总表.xlsx'


classifyJson = {
    "color": {
        "A": "4EAC5B",
        "B": "A0CD62",
        "C": "FFFD55",
        "D": "F7C143",
        "absence": "EA3323",
        "leave": "BFBFBF"
    },
    "蝙蝠": {
        "A": 390,
        "B": 385,
        "C": 380,
        "max": 401
    },
    "西瓜": {
        "A": 317,
        "B": 314,
        "C": 310,
        "max":317
    },
    "金币": {
        "A": 3090,
        "B": 3080,
        "C": 3070,
        "max":3110
    },
    "飞镖": {
        "A": 600,
        "B": 598,
        "C": 596,
        "max": 601,
    },
    "礼带": {
        "A": 76,
        "B": 73,
        "C": 70,
        "max": 82
    },
    "河豚": {
        "A": 185,
        "B": 180,
        "C": 175,
        "max": 197
    },
    "宝箱": {
        "A": 60,
        "B": 58,
        "C": 57,
        "max": 62
    },
    "无尽": {
        "A": 4067,
        "B": 4066,
        "C": 4065,
        "max": 4067
    }
}


def get_file_content(file):
    try:
        with open(file, 'rb') as f:
            image_data = f.read()
        image_b64 = base64.b64encode(image_data)
        return image_b64
    except Exception as e:
        raise Exception('')

    
def getOCRResult(image_path, templateSign):
    try:
        recognize_api_url = "https://aip.baidubce.com/rest/2.0/solution/v1/iocr/recognise"
        access_token = '24.124e94e468e569d107c6d5d4a95e951b.2592000.1586941045.282335-18723556'
        image = get_file_content(image_path)
        print('识别中...', image_path)
        recognize_body = "access_token=" + access_token + "&templateSign=" + templateSign + "&image=" + quote(image)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded",
            'charset': "utf-8"
        }
        response = requests.post(recognize_api_url, data=recognize_body, headers=headers)
        responseJson = json.loads(response.text)
        result = responseJson['data']
        if not result['isStructured']:
            print('识别失败，请确认识别是否完整，具体规则可联系管理员获知')
            exit(2)
        return result['ret']

    except Exception as e:
        raise Exception('')


def getImgInfo(signId):
    print('图片格式必须为: pic+数字，请务必修改')
    photoType = input("请输入图片后缀 (默认为: png):")
    if photoType == '': photoType = 'png'

    result = {}
    i = 1
    while 1:
        imagepath = os.path.dirname(os.path.realpath(sys.argv[0])) + '/' + 'pic' + str(i) + '.' + photoType

        try:
            rawResult = getOCRResult(imagepath, templateSignList[signId])
            i += 1
        except Exception as e:
            if i == 1:
                print('未找到文件，请确认文件名是否正确')
                exit(1)
            else:
                break
        if signId == 1:
            result = {**result, **generatePreparationDungeonData(rawResult)}
        elif signId == 2:
            result = {**result, **generateAbyssData(rawResult)}
        elif signId == 3:
            result = {**result, **generateFamilyWarDungeonData(rawResult)}
        else:
            result = {**result, **generateFightData(rawResult)}

    print('数据如下：')
    print(result)
    return result


def getNameList():
    global excelPath
    inputPath = input('请输入表格名 (默认为: 胧月成绩统计总表.xlsx):')
    if inputPath != '': excelPath = inputPath
    excelPath = os.path.dirname(os.path.realpath(sys.argv[0])) + '/' + excelPath
    excel = load_workbook(excelPath)
    FuBen = excel['副本']
    nameCells = FuBen['B']
    for cell in nameCells :
        if (cell.value != 'id') and (cell.value != None) and cell.fill.start_color.index != 1 :
            nameList.append(str(cell.value))


def generatePreparationDungeonData(rawData):
    j = 0
    result = {}
    while j < len(rawData):
        if 'member' in rawData[j]['word_name'] and 'score' in rawData[j+1]['word_name']:
            name = rawData[j]['word']
            score = rawData[j+1]['word']
            if "新人" in name or "族员" in name or "精英" in name or "豪杰" in name or "长老" in name or "副族长" in name or "族长" in name:
                name = name.replace('新人','')
                name =name.replace('族员','')
                name =name.replace('精英','')
                name =name.replace('豪杰','')
                name =name.replace('族长','')
                name =name.replace('副族长','')
            result.update({name: score})
        else:
            print('数据出错，请重试。')
        j += 2
    return result


def generateAbyssData(rawData):
    j = 0
    result = {}
    while j < len(rawData):
        if 'member' in rawData[j]['word_name'] and 'damage' in rawData[j+1]['word_name'] and 'times' in rawData[j+2]['word_name']:
            name = rawData[j]['word']
            if "新人" in name or "族员" in name or "精英" in name or "豪杰" in name or "副族长" in name or "族长" in name:
                name = name[0:len(name) - 2]
            # name = rawData[j]['word'][0:len(rawData[j]['word']) - 2]
            damage = rawData[j+1]['word']
            times = rawData[j+2]['word']
            try:
                average = str(int(int(damage) / int(times)))
            except Exception as e:
                print('\033[0;37;41mERROR\033[0m 族员:%s的平均伤害计算失败,请留意是否识别出错。他的成绩为:%s,%s' % (name, damage, times))
                average = '0'
            
            result.update({name: [damage, times, average]})
        else:
            print('数据出错，请重试。')
        j += 3
    return result


def generateFamilyWarDungeonData(rawData):
    j = 0
    tempDict = {}
    result = {}
    while j < len(rawData):
        tempDict.update({rawData[j]['word_name']:rawData[j]['word']})
        j += 1
    j = 1
    while j <= 6:
        name = 'member#' + str(j)
        score = 'score#' + str(j)
        result.update({tempDict[name]:tempDict[score]})
        j += 1
    return result


def generateFightData(rawData):
    return


def writeDungeonData(OCRResult, dungeonType):
    xlsx = load_workbook(excelPath)
    sheet = xlsx['副本']

    col = sheet.max_column + 1
    inputCol = input('将在表格第%s列插入数据。输入数字更改，或enter确认:' %(col))
    if inputCol != '' : col = int(inputCol)
    maxScore = int(classifyJson[dungeonList[dungeonType]]['max'])
    count = 0
    preScore = {'score':maxScore}
    nextScore = {'score':maxScore}
    currentScore = {'score':maxScore}
    currentRow = 0
    nextRow = 0
    for name, score in OCRResult.items():
        # print(preScore,currentScore,nextScore)
        row = nameList.index(name) if (name in nameList) else -1
        if row < 0 :
            maxSame = 0
            for i in range(len(nameList) - 1) :
                currentSame = fuzz.partial_ratio(nameList[i], name)
                if currentSame > maxSame :
                    maxSame = currentSame
                    row = i
            if row < 0 :
                print("\033[0;37;41mERROR\033[0m 没有找到族员:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认他是否改名。不然就是程序出错了诶嘿😛" % (name,score,count+1))
                continue
            elif  maxSame < 0.5:
                print("\033[0;30;43mWARN\033[0m 没有找到族员:\033[0;30;47m%s\033[0m，名字最接近的族员是:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s\033[0m。请留意匹配是否出错。" % (name,nameList[row],score))
        try:
            if 'score' in currentScore and 'score' in preScore and 'score' in nextScore and (float(currentScore['score']) > float(preScore['score']) or float(currentScore['score']) < float(nextScore['score']) or float(currentScore['score']) / float(preScore['score']) < 0.5 ):
                if 'name' in currentScore:
                    print("\033[0;30;43mINFO\033[0m 族员:\033[0;30;47m%s\033[0m的成绩较为异常。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认是否识别有误。(not the right rank)" % (nameList[currentRow],currentScore['score'],count-1))
            else:
                preScore = currentScore
            currentScore = nextScore
            currentRow = nextRow
            nextScore = {'name':name,'score':score}
            nextRow = row
        except Exception:
            pass

        score = 0
        row = 0
        try:
            if 'score' in currentScore:
                score = float(currentScore['score'])
                row = currentRow + 3
            elif  'score' in nextScore:
                score = float(nextScore['score'])
                row = nextRow + 3
            else:
                print("\033[0;37;41mERROR\033[0m。程序出错了。(get current score or next score error)")
        except Exception:
            if 'score' in currentScore:
                score = currentScore['score']
                row = currentRow + 3
            elif 'score' in nextScore:
                score = nextScore['score']
                row = nextRow + 3
            else:
                print("\033[0;37;41mERROR\033[0m。程序出错了。(get current score or next score error)")

        sheet.cell(row,col).value = score
        count += 1
    #将最后一个加进去
    try:
        if float(nextScore['score']) / float(currentScore['score']) < 0.5:
            print("\033[0;30;43mINFO\033[0m 族员:\033[0;30;47m%s\033[0m的成绩较为异常。他的成绩是:\033[0;37;44m%s\033[0m。他的大概排名为:\033[0;37;44m%d\033[0m。请确认是否识别有误。(not the right rank)" % (nameList[nextRow],nextScore['score'],count))
        score = float(nextScore['score'])
    except Exception:
        score = nextScore['score']
    
    row = nextRow + 3
    sheet.cell(row,col).value = score
        
    xlsx.save(excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 成功录入%d条数据！" % (count))
    return col


def writeAbyssData(OCRResult):
    xlsx = load_workbook(excelPath)
    sheet = xlsx['深渊']

    col = sheet.max_column + 1
    inputCol = input('将在表格第%s,%s,%s列插入数据。输入数字更改，或enter确认:' %(col, col + 1, col + 2))
    if inputCol != '' : col = int(inputCol)

    count = 0
    for name, data in OCRResult.items() :
        damage = data[0]
        times = data[1]
        average = data[2]
        row = nameList.index(name) if (name in nameList) else -1
        if row < 0 :
            maxSame = 0
            for i in range(len(nameList) - 1) :
                currentSame = fuzz.partial_ratio(nameList[i], name)
                if currentSame > maxSame :
                    maxSame = currentSame
                    row = i
            if row < 0 :
                print("\033[0;37;41mERROR\033[0m 没有找到族员:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s,%s,%s\033[0m。请确认他是否改名。不然就是程序出错了诶嘿😛" % (name, damage, times, average))
                continue
            else :
                print("\033[0;30;43mWARN\033[0m 没有找到族员:\033[0;30;47m%s\033[0m，名字最接近的族员是:\033[0;30;47m%s\033[0m。他的成绩是:\033[0;37;44m%s,%s,%s\033[0m。请留意匹配是否出错。" % (name,nameList[row],damage, times, average))
                        
        row += 3
        try:
            sheet.cell(row,col).value = float(damage)
            sheet.cell(row,col+1).value = float(times)
            sheet.cell(row,col+2).value = float(average)
        except Exception as e:
            print('\033[0;37;41mERROR\033[0m \033[0;30;47m%s\033[0m的成绩登记有误，请留意是否出错，他的成绩为:%s,%s,%s' % (name, damage, times, average))
            sheet.cell(row,col).value = damage
            sheet.cell(row,col+1).value = times
            sheet.cell(row,col+2).value = average

        count += 1
        
    xlsx.save(excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 成功录入%d条数据！" % (count))
    return col


def writeFightData(OCRResult):
    print('还没写……')
    return


def decorateDungeonData(col, dungeonType):
    input("输入任意键开始填充颜色……")
    xlsx = load_workbook(excelPath)
    FuBen = xlsx['副本']
    classify = classifyJson[dungeonList[dungeonType]]
    colorList = classifyJson['color']
    FuBen.cell(2, col).value = dungeonList[dungeonType]

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin',color='000000'), bottom=Side(border_style='thin',color='000000'))
    
    defaultAbsence = False
    for i in range(3, len(nameList)+ 3): 
        color = ""
        if FuBen.cell(i,col).value == None:
            if not defaultAbsence:
                noScore = input("%s没有成绩，是否请假？(默认否,全部未请假输入all)[N/y/all]:" %(FuBen.cell(i,2).value))
                if noScore == '': noScore = "N"
                while noScore != 'N' and noScore != 'n' and noScore != 'Y' and noScore != 'y' and noScore != 'all':
                    noScore = input("输入无效，请输入‘y’、‘n’、‘all’或直接enter:")
                    if noScore == '': noScore = "N"
                if noScore == 'all': 
                    defaultAbsence = True
                    noScore = 'N'
            else:
                noScore = 'N'       
            color = colorList['absence'] if noScore == "N" or noScore == "n" else colorList['leave']
        elif int(FuBen.cell(i,col).value) >= classify['A']:
            color = colorList['A']
        elif int(FuBen.cell(i,col).value) < classify['A'] and int(FuBen.cell(i,col).value) >= classify['B']:
            color = colorList['B']
        elif  int(FuBen.cell(i,col).value) < classify['B'] and int(FuBen.cell(i,col).value) >= classify['C']:
            color = colorList['C']
        else:
            color = colorList['D']
        FuBen.cell(i, col).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
        FuBen.cell(i, col).border = border
    
    xlsx.save(excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 颜色填充成功！")


def decorateAbyssData(col):
    xlsx = load_workbook(excelPath)
    sheet = xlsx['深渊']

    absenceColor = 'EA3323'
    leaveColor = 'BFBFBF'

    border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin',color='000000'), bottom=Side(border_style='thin',color='000000'))

    defaultAbsence = False
    for i in range(3, len(nameList)+ 3): 
        color = ""
        if sheet.cell(i,col).value == None:
            if not defaultAbsence:
                noScore = input("%s没有成绩，是否请假？(默认否,全部未请假输入all)[N/y/all]:" %(sheet.cell(i,2).value))
                if noScore == '': noScore = "N"
                while noScore != 'N' and noScore != 'n' and noScore != 'Y' and noScore != 'y' and noScore != 'all':
                    noScore = input("输入无效，请输入‘y’、‘n’、‘all’或直接enter:")
                    if noScore == '': noScore = "N"
                if noScore == 'all': 
                    defaultAbsence = True
                    noScore = 'N'
            else:
                noScore = 'N'       
            color = absenceColor if noScore == "N" or noScore == "n" else leaveColor

            sheet.cell(i, col).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
            sheet.cell(i, col+1).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
            sheet.cell(i, col+2).fill = PatternFill(fill_type = 'solid', start_color=color, end_color=color)
        
        #描边
        sheet.cell(i, col).border = border
        sheet.cell(i, col+1).border = border
        sheet.cell(i, col+2).border = border

    xlsx.save(excelPath)
    print("\033[0;30;42mSUCCESS\033[0m 颜色填充成功！")


def processData(result, sheetId, **option):
    getNameList()
    if sheetId == 1 or sheetId == 3:
        col = writeDungeonData(result, option['dungeonType'])
        decorateDungeonData(col, option['dungeonType'])
    elif sheetId == 2:
        col = writeAbyssData(result)
        decorateAbyssData(col)
    else:
        writeFightData(result)


def dungeonPreparation():
    dungeonType = input('请选择副本类型 (默认为: 礼带)\n1.蝙蝠 2.西瓜 3.金币 4.飞镖 5.礼带 6.河豚 7.宝箱 8.无尽\n:')
    if dungeonType == '': dungeonType = '5'
    result = getImgInfo(1)
    processData(result, 1, dungeonType = dungeonType)


def abyss():
    result = getImgInfo(2)
    processData(result, 2)
    return


def familyWarDungeon():
    dungeonType = input('请选择副本类型 (默认为: 礼带)\n1.蝙蝠 2.西瓜 3.金币 4.飞镖 5.礼带 6.河豚 7.宝箱 8.无尽\n:')
    if dungeonType == '': dungeonType = '5'
    result = getImgInfo(3)
    processData(result, 3, dungeonType = dungeonType)


def familyWarFight():
    return


if __name__ == '__main__' :
    job = input("请选择要登记的成绩\n1.备战副本\n2.深渊\n3.家族战副本\n4.家族战33\n(1/2/3/4):")
    while 1:
        if job == '1':
            dungeonPreparation()
            break
        elif job == '2':
            abyss()
            break
        elif job == '3':
            familyWarDungeon()
            break
        elif job == '4':
            familyWarFight()
            break
        else:
            print('不合法输入，请重新输入。')
            job = input("请选择要登记的成绩\n1.备战副本\n2.深渊\n3.家族战副本\n4.家族战33\n(1/2/3/4):")
    input("登记完毕。按任意键结束程序……")
